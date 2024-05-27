import pdfplumber
import os, openpyxl
from pathlib import Path
from openpyxl import load_workbook

# Get downloads path
def get_directory():
    path = Path()
    download_path = str(path.home() / 'Downloads')
    return download_path + '/'

# Get specific file in downloads directory
def get_file(directory):
    files = os.listdir(directory)
    for f in files:
        try:
            if 'Nubank_' in f:  # Search file's name
                return f
        except FileNotFoundError:
            print(FileNotFoundError)

# The data is located at the end of pdf (page 4 -> index 3)
def open_pdf(file):
    with pdfplumber.open(file) as pdf:
        page = pdf.pages[3]  # data location
        table = page.extract_table()
        organize_list(table)
        return table  # just for pytest


# removing empty strings on list
def organize_list(data):
    for i in range(len(data)):
        data[i].pop(0)
        data[i].remove('')
    py_excel(data)

def py_excel(d):

    # create new workbook
    wb = openpyxl.Workbook()

    '''
        if you want to load an existing file
        wb = load_workbook('woksheet.xlsx')   worksheet wanted
    '''

    '''
        creating sheet page
        wb.create_sheet('Nubank Fatura')

        access sheet page
        fatura_page = wb['Nubank Fatura'] 
    '''
    
    fatura_page = wb.create_sheet('Credit_expenses')

    # Columns
    fatura_page.append(['Name', 'Amount']) 
    
    # insert values in each cell
    for i, row in enumerate(d):
        for j, value in enumerate(row):
            fatura_page.cell(row=i + 2, column=j + 1, value=value)

    # printing values in console
    for row in fatura_page.iter_rows(min_row=2, max_row=len(d), max_col=2, values_only=True):
        print(row)
    for row in fatura_page.iter_rows(min_row=2, max_row=len(d), max_col=2, values_only=False):
        print(row)

    # save modifications
    wb.save('New_worksheet.xlsx')

def main():
    d = get_directory()
    user = get_file(d)
    open_pdf(d+user)


if __name__ == "__main__":
    main()
